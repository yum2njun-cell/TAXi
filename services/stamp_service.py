"""
TAXi - 인지세 관리 서비스
"""

import pandas as pd
import numpy as np
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import io
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.activity_logger import log_stamp_tax_activity


class StampTaxService:
    """인지세 관리 서비스"""
    
    # 컬럼 한글명
    COL_OPP_ID = "사업기회번호"
    COL_VALUES = [
        "인건비탭 전체 합계",
        "외주비탭 전체 합계",
        "HW 영역 합계",
        "공사_공사 품목 합계",
        "공사MA탭 전체 합계",
        "경비탭 전체 합계",
    ]
    
    # Excel 회계 형식
    ACCOUNTING_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
    
    # 배경색 (HEX 6자리)
    COND1_COLOR = "DAF2D0"  # 조건1
    COND2_COLOR = "F2CEEF"  # 조건2
    COND3_COLOR = "FBE2D5"  # 조건3
    
    def __init__(self, output_dir: str = "data/outputs/stamp_tax"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def load_raw_file(self, file_path: Path, sheet_name=0) -> pd.DataFrame:
        """Raw 파일 로드"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            required = [self.COL_OPP_ID] + self.COL_VALUES
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise ValueError(f"입력 파일에 다음 컬럼이 없습니다: {missing}")
            
            # 숫자 변환
            for c in self.COL_VALUES:
                df[c] = pd.to_numeric(
                    df[c].astype(str).str.replace(",", "", regex=False),
                    errors="coerce"
                )
            
            df = df.replace({np.inf: np.nan, -np.inf: np.nan})
            
            log_stamp_tax_activity(
                f"Raw 파일 로드 완료",
                {"file": file_path.name, "rows": len(df)}
            )
            
            return df
            
        except Exception as e:
            log_stamp_tax_activity(
                f"Raw 파일 로드 실패: {str(e)}",
                {"file": file_path.name}
            )
            raise e
    
    def load_total_file(self, file_path: Path, sheet=0) -> Dict[str, str]:
        """Total 파일에서 확인 컬럼 매칭 딕셔너리 생성"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet, dtype=str)
            cols = {c.strip() for c in df.columns.astype(str)}
            
            if ('사업기회번호' not in cols) or ('확인' not in cols):
                raise ValueError("total 파일에 '사업기회번호'(C열), '확인'(B열) 헤더가 있어야 합니다.")
            
            df['사업기회번호'] = df['사업기회번호'].astype(str).str.strip()
            df['확인'] = df['확인'].astype(str).fillna('').str.strip()
            
            last_df = (
                df.dropna(subset=['사업기회번호'])
                  .drop_duplicates(subset=['사업기회번호'], keep='last')
            )
            
            lookup = dict(zip(last_df['사업기회번호'], last_df['확인']))
            
            log_stamp_tax_activity(
                f"Total 파일 로드 완료",
                {"file": file_path.name, "records": len(lookup)}
            )
            
            return lookup
            
        except Exception as e:
            log_stamp_tax_activity(
                f"Total 파일 로드 실패: {str(e)}",
                {"file": file_path.name}
            )
            raise e
    
    def build_pivot(self, df: pd.DataFrame, fill_zero: bool = False) -> pd.DataFrame:
        """피벗 테이블 생성"""
        pivot = pd.pivot_table(
            df,
            index=self.COL_OPP_ID,
            values=self.COL_VALUES,
            aggfunc="sum",
            dropna=False,
        )
        
        desired_order = [
            "인건비탭 전체 합계",
            "외주비탭 전체 합계",
            "HW 영역 합계",
            "공사_공사 품목 합계",
            "공사MA탭 전체 합계",
            "경비탭 전체 합계",
        ]
        pivot = pivot.reindex(columns=desired_order)
        
        if fill_zero:
            pivot = pivot.fillna(0)
        
        return pivot
    
    def categorize_data(self, pivot: pd.DataFrame) -> Tuple[Dict, pd.Series, pd.Index, pd.Index, pd.Index]:
        """데이터 조건별 분류"""
        p = pivot.copy()
        
        # 조건 마스크
        c1 = (p["공사_공사 품목 합계"].fillna(0) > 0) | (p["공사MA탭 전체 합계"].fillna(0) > 0)
        c2 = (p["HW 영역 합계"].fillna(0) > 0) & (
             (p["인건비탭 전체 합계"].fillna(0) > 0) | (p["외주비탭 전체 합계"].fillna(0) > 0)
        )
        c3 = (p["경비탭 전체 합계"].fillna(0) > 0)
        
        category = np.where(c1, "조건1(공사/공사MA)",
                     np.where(c2, "조건2(HW + 인건비/외주)",
                     np.where(c3, "조건3(경비)", "해당없음")))
        
        cat_map = {str(idx): cat for idx, cat in zip(p.index, category)}
        
        c1_ids = p.index[c1]
        c2_ids = p.index[~c1 & c2]
        c3_ids = p.index[~c1 & ~c2 & c3]
        
        return cat_map, category, c1_ids, c2_ids, c3_ids
    
    def process_files(self, raw_file, total_file=None, fill_zero: bool = False) -> Tuple[pd.DataFrame, pd.DataFrame, bytes]:
        """인지세 파일 처리 메인 로직"""
        try:
            log_stamp_tax_activity("인지세 파일 처리 시작", {})
            
            # Raw 파일 로드
            if isinstance(raw_file, (str, Path)):
                raw_path = Path(raw_file)
            else:
                # UploadedFile 객체인 경우
                raw_path = Path(f"temp_{raw_file.name}")
                with open(raw_path, 'wb') as f:
                    f.write(raw_file.read())
            
            df_raw = self.load_raw_file(raw_path)
            
            # Total 파일 로드 (선택)
            total_lookup = None
            if total_file:
                if isinstance(total_file, (str, Path)):
                    total_path = Path(total_file)
                else:
                    total_path = Path(f"temp_{total_file.name}")
                    with open(total_path, 'wb') as f:
                        f.write(total_file.read())
                
                total_lookup = self.load_total_file(total_path)
            
            # 피벗 테이블 생성
            df_pivot = self.build_pivot(df_raw, fill_zero=fill_zero)
            
            # 엑셀 생성
            excel_data = self._generate_excel(df_raw, df_pivot, total_lookup)
            
            log_stamp_tax_activity(
                "인지세 파일 처리 완료",
                {
                    "raw_rows": len(df_raw),
                    "pivot_rows": len(df_pivot),
                    "has_total": total_lookup is not None
                }
            )
            
            return df_raw, df_pivot, excel_data
            
        except Exception as e:
            log_stamp_tax_activity(
                f"인지세 파일 처리 실패: {str(e)}",
                {"error_type": "processing_error"}
            )
            raise e
    
    def _generate_excel(self, df_raw: pd.DataFrame, df_pivot: pd.DataFrame, 
                       total_lookup: Optional[Dict] = None) -> bytes:
        """엑셀 파일 생성"""
        try:
            excel_buffer = io.BytesIO()
            
            # 조건별 분류
            cat_map, category, c1_ids, c2_ids, c3_ids = self.categorize_data(df_pivot)
            
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                # 시트 작성
                df_raw.to_excel(writer, sheet_name="Raw", index=False)
                df_pivot.to_excel(writer, sheet_name="Pivot")
                
                summary = df_raw.groupby(self.COL_OPP_ID, dropna=False)[self.COL_VALUES].sum()
                summary.to_excel(writer, sheet_name="Summary")
                
                # Extract 시트
                self._write_extract_sheet(writer, c1_ids, c2_ids, c3_ids, category, df_pivot)
                
                # 서식 적용
                self._apply_excel_formatting(writer, df_pivot, cat_map, total_lookup)
            
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except Exception as e:
            raise e
    
    def _write_extract_sheet(self, writer, c1_ids, c2_ids, c3_ids, category, pivot):
        """Extract 시트 작성"""
        extract_ws_name = "Extract"
        extract_blocks = [
            ("조건1 목록(공사/공사MA)", pd.DataFrame({self.COL_OPP_ID: c1_ids})),
            ("조건2 목록(HW + 인건비/외주)", pd.DataFrame({self.COL_OPP_ID: c2_ids})),
            ("조건3 목록(경비)", pd.DataFrame({self.COL_OPP_ID: c3_ids})),
        ]
        
        df_flags = pd.DataFrame({
            self.COL_OPP_ID: pivot.index, 
            "분류(우선순위적용)": category
        })
        
        start_row = 1
        for title, df_block in extract_blocks:
            pd.DataFrame({title: []}).to_excel(
                writer, sheet_name=extract_ws_name, 
                index=False, startrow=start_row-1
            )
            df_block.to_excel(
                writer, sheet_name=extract_ws_name, 
                index=False, startrow=start_row
            )
            start_row += max(2, len(df_block) + 3)
        
        df_flags.to_excel(
            writer, sheet_name=extract_ws_name, 
            index=False, startrow=start_row
        )
    
    def _apply_excel_formatting(self, writer, df_pivot, cat_map, total_lookup):
        """엑셀 서식 적용"""
        wb = writer.book
        ws = writer.sheets["Pivot"]
        
        # 회계 스타일
        acct_style = NamedStyle(name="acct")
        acct_style.number_format = self.ACCOUNTING_FORMAT
        try:
            wb.add_named_style(acct_style)
        except ValueError:
            pass
        
        # 값 영역 회계 서식
        min_row, min_col = 2, 2
        max_row = ws.max_row
        max_col = 1 + len(df_pivot.columns)
        
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for cell in row:
                cell.style = "acct"
        
        # 배경색
        fill1 = PatternFill(start_color=self.COND1_COLOR, end_color=self.COND1_COLOR, fill_type="solid")
        fill2 = PatternFill(start_color=self.COND2_COLOR, end_color=self.COND2_COLOR, fill_type="solid")
        fill3 = PatternFill(start_color=self.COND3_COLOR, end_color=self.COND3_COLOR, fill_type="solid")
        
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=1):
            for cell in row:
                key = str(cell.value)
                cat = cat_map.get(key)
                if not cat:
                    continue
                if cat.startswith("조건1"):
                    cell.fill = fill1
                elif cat.startswith("조건2"):
                    cell.fill = fill2
                elif cat.startswith("조건3"):
                    cell.fill = fill3
        
        # 열 너비
        ws.column_dimensions["A"].width = 22
        for j in range(min_col, max_col + 1):
            ws.column_dimensions[get_column_letter(j)].width = 18
        
        # total 매칭
        if total_lookup:
            if not ws.cell(row=1, column=8).value:
                ws.cell(row=1, column=8, value="확인(total)")
            
            for r in range(2, max_row + 1):
                a_cell = ws.cell(row=r, column=1)
                h_cell = ws.cell(row=r, column=8)
                
                if getattr(a_cell.fill, "fill_type", None) != "solid":
                    h_cell.value = ""
                    continue
                
                key = "" if a_cell.value is None else str(a_cell.value).strip()
                h_cell.value = total_lookup.get(key, "")
        
        # 공통 서식
        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin = Side(style="thin")
        
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                              min_col=1, max_col=ws.max_column):
            for cell in r:
                is_header = (cell.row == 1)
                cell.font = Font(name="맑은 고딕", size=10, bold=is_header)
                cell.alignment = center
        
        # H1 테두리
        h1 = ws.cell(row=1, column=8)
        h1.font = Font(name="맑은 고딕", size=10, bold=True)
        h1.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # 다른 시트들도 서식 적용
        for name in ["Raw", "Summary", "Extract"]:
            ws_other = wb[name]
            for r in ws_other.iter_rows(min_row=1, max_row=ws_other.max_row,
                                       min_col=1, max_col=ws_other.max_column):
                for cell in r:
                    is_header = (cell.row == 1)
                    cell.font = Font(name="맑은 고딕", size=10, bold=is_header)
                    cell.alignment = center


# 싱글톤 인스턴스
_stamp_service = None

def get_stamp_tax_service() -> StampTaxService:
    """인지세 서비스 인스턴스 반환"""
    global _stamp_service
    if _stamp_service is None:
        _stamp_service = StampTaxService()
    return _stamp_service