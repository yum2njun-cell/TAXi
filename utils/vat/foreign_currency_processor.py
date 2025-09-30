# utils/vat/foreign_currency_processor.py
"""외화획득명세서 데이터 처리기 (TAXi 시스템 통합버전)"""

import pandas as pd
import numpy as np
import io
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, timedelta
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ForeignCurrencyProcessor:
    """외화획득명세서 통합 데이터 처리기"""
    
    def __init__(self):
        self.year = None
        self.quarter = None
        self.raw_data = {}
        self.processed_data = {}
        
    def set_period(self, year: int, quarter: int):
        """처리 기간 설정"""
        self.year = year
        self.quarter = quarter
    
    def get_quarter_dates(self) -> Tuple[datetime, datetime]:
        """분기 날짜 계산"""
        quarter_months = {
            1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)
        }
        
        start_month, end_month = quarter_months[self.quarter]
        start_date = datetime(self.year, start_month, 1)
        
        if end_month == 12:
            end_date = datetime(self.year, 12, 31)
        else:
            end_date = datetime(self.year, end_month + 1, 1) - timedelta(days=1)
        
        return start_date, end_date
    
    def process_export_data(self, files: List[Any]) -> Dict[str, Any]:
        """수출이행내역 파일들 취합 처리"""
        try:
            export_columns = [
                '신고번호', '수출종류', '송품장번호', '수리일자', '선적기한',
                '신고수량', '신고수량단위', '선적수량', '선적수량단위',
                '신고중량', '선적중량', '포장여부', '선적여부', '선적일자',
                '통화부호', '환율', '수출금액(미화)', '수출금액(원화)',
                'B/L', '수출자구분', '거래구분', '결제방법', '결재금액',
                '인도조건', '운송수단코드명', '구매자상호', '구매자부호',
                '운임', '보험료', '신용장번호'
            ]
            
            all_dataframes = []
            file_logs = []
            
            for uploaded_file in files:
                try:
                    df = pd.read_excel(uploaded_file, sheet_name=0)
                    df.columns = [col.strip() for col in df.columns]
                    df = df.dropna(how='all')
                    
                    selected_df = pd.DataFrame()
                    for col in export_columns:
                        if col in df.columns:
                            selected_df[col] = df[col]
                        else:
                            selected_df[col] = None
                    
                    all_dataframes.append(selected_df)
                    file_logs.append({
                        'file': uploaded_file.name,
                        'rows': len(df),
                        'status': 'success'
                    })
                except Exception as e:
                    file_logs.append({
                        'file': uploaded_file.name,
                        'rows': 0,
                        'status': f'error: {str(e)}'
                    })
            
            if all_dataframes:
                consolidated_data = pd.concat(all_dataframes, ignore_index=True)
            else:
                consolidated_data = pd.DataFrame(columns=export_columns)
            
            excel_data = self._create_excel_with_logs(
                consolidated_data, file_logs, 
                data_sheet_name='수출내역',
                log_sheet_name='취합로그'
            )
            
            return {
                'data': consolidated_data,
                'excel_data': excel_data,
                'logs': file_logs,
                'summary': {
                    'total_rows': len(consolidated_data),
                    'total_files': len(file_logs),
                    'success_files': sum(1 for log in file_logs if log['status'] == 'success')
                }
            }
            
        except Exception as e:
            return {'error': str(e), 'data': pd.DataFrame()}
    
    def process_exchange_data(self, files: List[Any]) -> Dict[str, Any]:
        """환율 파일들 취합 처리"""
        try:
            start_date, end_date = self.get_quarter_dates()
            currency_data = {}
            file_logs = []
            
            for uploaded_file in files:
                try:
                    xl_file = pd.ExcelFile(uploaded_file)
                    sheet_names = xl_file.sheet_names
                    
                    if not sheet_names:
                        continue
                    
                    currency_code = sheet_names[0].upper().strip()
                    df = pd.read_excel(uploaded_file, sheet_name=sheet_names[0])
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    # 일자와 환율 컬럼 찾기
                    date_col = None
                    rate_col = None
                    
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(keyword in col_lower for keyword in ['일자', 'date', '날짜', '기준일']):
                            date_col = col
                        elif any(keyword in col_lower for keyword in ['환율', 'rate', '율']):
                            rate_col = col
                    
                    if date_col is None and len(df.columns) >= 1:
                        date_col = df.columns[0]
                    if rate_col is None and len(df.columns) >= 2:
                        rate_col = df.columns[1]
                    
                    if date_col is None or rate_col is None:
                        file_logs.append({
                            'file': getattr(uploaded_file, 'name', 'unknown'),
                            'currency': currency_code,
                            'rows': 0,
                            'status': f'error: 필수 컬럼 없음'
                        })
                        continue
                    
                    df_clean = df[[date_col, rate_col]].copy()
                    df_clean.columns = ['일자', '환율']
                    
                    df_clean['일자'] = pd.to_datetime(df_clean['일자'], errors='coerce')
                    df_clean = df_clean.dropna(subset=['일자'])
                    df_clean['환율'] = pd.to_numeric(df_clean['환율'], errors='coerce')
                    df_clean = df_clean.dropna(subset=['환율'])
                    
                    # JPY 특별 처리
                    if currency_code == 'JPY':
                        df_clean['환율'] = df_clean['환율'] / 100
                    
                    # 분기 필터링
                    buffer_start = start_date - timedelta(days=10)
                    mask = (df_clean['일자'] >= buffer_start) & (df_clean['일자'] <= end_date)
                    df_filtered = df_clean[mask].copy()
                    
                    df_filtered = df_filtered.sort_values('일자')
                    df_filtered = df_filtered.drop_duplicates(subset=['일자'], keep='last')
                    
                    currency_data[currency_code] = df_filtered
                    file_logs.append({
                        'file': getattr(uploaded_file, 'name', 'unknown'),
                        'currency': currency_code,
                        'rows': len(df_filtered),
                        'status': 'success'
                    })
                    
                except Exception as e:
                    file_logs.append({
                        'file': getattr(uploaded_file, 'name', 'unknown'),
                        'currency': 'unknown',
                        'rows': 0,
                        'status': f'error: {str(e)}'
                    })
            
            if not currency_data:
                return {'error': '유효한 환율 데이터가 없습니다', 'data': pd.DataFrame()}
            
            # 분기의 모든 날짜 생성
            all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
            consolidated_data = pd.DataFrame({'일자': all_dates})
            
            # 각 통화별로 환율 데이터 병합
            for currency_code, currency_df in currency_data.items():
                consolidated_data = pd.merge(
                    consolidated_data,
                    currency_df.rename(columns={'환율': currency_code}),
                    on='일자',
                    how='left'
                )
                consolidated_data[currency_code] = consolidated_data[currency_code].ffill()
            
            excel_data = self._create_excel_with_logs(
                consolidated_data, file_logs,
                data_sheet_name='환율',
                log_sheet_name='처리로그'
            )
            
            return {
                'data': consolidated_data,
                'excel_data': excel_data,
                'logs': file_logs,
                'summary': {
                    'total_days': len(consolidated_data),
                    'currencies': list(currency_data.keys()),
                    'currency_count': len(currency_data)
                }
            }
            
        except Exception as e:
            return {'error': str(e), 'data': pd.DataFrame()}


    def process_invoice_data(self, invoice_files: List[Any], export_data: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
        """인보이스 발행내역 처리"""
        try:
            required_columns = [
                '기안자', '인보이스 넘버', 'PJT#', '전표번호', 
                'Customer', '국가', 'Doc. Date', '금액', '단위', '구분', '비고'
            ]
            
            all_invoice_data = []
            file_logs = []
            
            for uploaded_file in invoice_files:
                try:
                    sheet_name = f"{str(self.year)[-2:]}년"
                    
                    try:
                        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                    except ValueError:
                        xl_file = pd.ExcelFile(uploaded_file)
                        df = pd.read_excel(uploaded_file, sheet_name=xl_file.sheet_names[0])
                    
                    df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
                    
                    # 필요한 컬럼만 선택
                    available_columns = []
                    for col in required_columns:
                        if col in df.columns:
                            available_columns.append(col)
                        elif col == 'Customer' and 'Custmer' in df.columns:
                            available_columns.append('Custmer')
                    
                    selected_df = df[available_columns].copy()
                    
                    if 'Custmer' in selected_df.columns:
                        selected_df = selected_df.rename(columns={'Custmer': 'Customer'})
                    
                    for col in required_columns:
                        if col not in selected_df.columns:
                            selected_df[col] = None
                    
                    selected_df = selected_df[required_columns]
                    
                    # 분기 필터링
                    if 'Doc. Date' in selected_df.columns:
                        selected_df['Doc. Date'] = pd.to_datetime(selected_df['Doc. Date'], errors='coerce')
                    
                    start_date, end_date = self.get_quarter_dates()
                    mask = (selected_df['Doc. Date'] >= start_date) & (selected_df['Doc. Date'] <= end_date)
                    filtered_df = selected_df[mask].copy()
                    
                    filtered_df['구분'] = '용역'
                    all_invoice_data.append(filtered_df)
                    
                    file_logs.append({
                        'file': uploaded_file.name,
                        'total_rows': len(selected_df),
                        'filtered_rows': len(filtered_df),
                        'status': 'success'
                    })
                    
                except Exception as e:
                    file_logs.append({
                        'file': uploaded_file.name,
                        'total_rows': 0,
                        'filtered_rows': 0,
                        'status': f'error: {str(e)}'
                    })
            
            # 인보이스 데이터 합치기
            if all_invoice_data:
                invoice_data = pd.concat(all_invoice_data, ignore_index=True)
            else:
                invoice_data = pd.DataFrame(columns=required_columns)
            
            # 수출내역 데이터 추가
            if export_data is not None and not export_data.empty:
                export_invoice_format = self._convert_export_to_invoice_format(export_data)
                consolidated_data = pd.concat([invoice_data, export_invoice_format], ignore_index=True)
            else:
                consolidated_data = invoice_data
            
            excel_data = self._create_excel_with_logs(
                consolidated_data, file_logs,
                data_sheet_name='인보이스 발행내역',
                log_sheet_name='처리로그'
            )
            
            return {
                'data': consolidated_data,
                'excel_data': excel_data,
                'logs': file_logs,
                'summary': {
                    'total_rows': len(consolidated_data),
                    'service_rows': len(consolidated_data[consolidated_data['구분'] == '용역']) if '구분' in consolidated_data.columns else 0,
                    'goods_rows': len(consolidated_data[consolidated_data['구분'] == '재화']) if '구분' in consolidated_data.columns else 0
                }
            }
            
        except Exception as e:
            return {'error': str(e), 'data': pd.DataFrame()}
    
    def process_a2_data(self, a2_files: List[Any], invoice_data: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
        """A2리스트 처리 및 정리"""
        try:
            if not a2_files:
                return {'error': 'A2 파일이 없습니다', 'data': pd.DataFrame()}
            
            uploaded_file = a2_files[0]
            df = pd.read_excel(uploaded_file, sheet_name=0)
            df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
            
            # G/L Account 컬럼으로 필터링
            gl_account_col_idx = self._letter_to_index('G')
            gl_account_col = df.columns[gl_account_col_idx]
            
            df[gl_account_col] = (df[gl_account_col]
                                .astype(str)
                                .str.strip()
                                .str.replace(',', '')
                                .str.replace('.0', '')
                                .str.replace('.00', ''))
            
            mask = df[gl_account_col].isin(['207101', '207102'])
            filtered_df = df[mask].copy()
            
            if filtered_df.empty:
                return {'error': '207101 또는 207102에 해당하는 데이터가 없습니다', 'data': pd.DataFrame()}
            
            processed_data = self._process_a2_structure(filtered_df, invoice_data)
            excel_data = self._create_a2_excel(processed_data)
            
            return {
                'data': processed_data,
                'excel_data': excel_data,
                'summary': {
                    'original_rows': len(df),
                    'filtered_rows': len(filtered_df),
                    'processed_rows': len(processed_data)
                }
            }
            
        except Exception as e:
            return {'error': str(e), 'data': pd.DataFrame()}
    
    def generate_final_report(self, invoice_data: pd.DataFrame, exchange_data: Optional[pd.DataFrame] = None, 
                            a2_data: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
        """최종 외화획득명세서 생성"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "외화획득명세서"
            
            self._create_final_report_headers(ws)
            self._fill_final_report_data(ws, invoice_data, exchange_data, a2_data)
            self._apply_final_report_styles(ws)
            self._add_final_report_formulas(ws)
            
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_data = excel_buffer.getvalue()
            
            return {
                'excel_data': excel_data,
                'summary': {
                    'total_rows': len(invoice_data) if invoice_data is not None else 0,
                    'currencies_used': list(exchange_data.columns[1:]) if exchange_data is not None else []
                }
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    # 헬퍼 메서드들
    def _convert_export_to_invoice_format(self, export_data: pd.DataFrame) -> pd.DataFrame:
        """수출내역 데이터를 인보이스 형식으로 변환"""
        required_columns = [
            '기안자', '인보이스 넘버', 'PJT#', '전표번호', 
            'Customer', '국가', 'Doc. Date', '금액', '단위', '구분', '비고'
        ]
        
        invoice_format_df = pd.DataFrame(columns=required_columns)
        
        invoice_format_df['기안자'] = None
        invoice_format_df['인보이스 넘버'] = export_data['신고번호'] if '신고번호' in export_data.columns else None
        invoice_format_df['PJT#'] = None
        invoice_format_df['전표번호'] = None
        invoice_format_df['Customer'] = export_data['구매자상호'] if '구매자상호' in export_data.columns else None
        invoice_format_df['국가'] = None
        invoice_format_df['Doc. Date'] = pd.to_datetime(export_data['선적일자'], errors='coerce') if '선적일자' in export_data.columns else None
        invoice_format_df['금액'] = export_data['수출금액(미화)'] if '수출금액(미화)' in export_data.columns else None
        invoice_format_df['단위'] = export_data['통화부호'] if '통화부호' in export_data.columns else None
        invoice_format_df['구분'] = '재화'
        invoice_format_df['비고'] = None
        
        return invoice_format_df
    
    def _letter_to_index(self, letter: str) -> int:
        """엑셀 열 문자를 인덱스로 변환"""
        result = 0
        for i, char in enumerate(reversed(letter.upper())):
            result += (ord(char) - ord('A') + 1) * (26 ** i)
        return result - 1
    
    def _process_a2_structure(self, filtered_data: pd.DataFrame, invoice_data: Optional[pd.DataFrame]) -> pd.DataFrame:
        """A2 데이터 구조 처리"""
        c_idx = self._letter_to_index('C')   # Doc.No.
        ac_idx = self._letter_to_index('AC')  # Doc.Date
        ai_idx = self._letter_to_index('AI')  # DR Amount(L)
        aj_idx = self._letter_to_index('AJ')  # CR Amount(L)
        
        # 원본 데이터
        col1_data = []
        for _, row in filtered_data.iterrows():
            doc_no = row.iloc[c_idx]
            doc_date = row.iloc[ac_idx]
            
            dr_amount = pd.to_numeric(row.iloc[ai_idx], errors='coerce')
            cr_amount = pd.to_numeric(row.iloc[aj_idx], errors='coerce')
            
            if pd.isna(dr_amount):
                dr_amount = 0
            if pd.isna(cr_amount):
                cr_amount = 0
            
            amount_diff = dr_amount - cr_amount
            
            col1_data.append({
                'Doc.No.': doc_no,
                'Amount': amount_diff,
                'Doc.Date': doc_date
            })
        
        df_col1 = pd.DataFrame(col1_data)
        
        # 중복 제거 및 집계
        grouped = df_col1.groupby('Doc.No.').agg({
            'Amount': 'sum',
            'Doc.Date': 'first'
        }).reset_index()
        grouped.columns = ['Doc.No.(중복제거)', '합계', 'Doc.Date']
        
        # 인보이스 매칭
        invoice_matched = []
        invoice_unmatched = []
        
        if invoice_data is not None and '전표번호' in invoice_data.columns:
            invoice_docs = invoice_data[['전표번호', '금액']].copy()
            invoice_docs['전표번호'] = invoice_docs['전표번호'].astype(str).str.strip()
            grouped['Doc.No.(중복제거)'] = grouped['Doc.No.(중복제거)'].astype(str).str.strip()
            
            for doc_no in grouped['Doc.No.(중복제거)']:
                matched = invoice_docs[invoice_docs['전표번호'] == doc_no]
                if not matched.empty:
                    for _, row in matched.iterrows():
                        invoice_matched.append({
                            '전표번호': row['전표번호'],
                            '금액': row['금액']
                        })
            
            matched_docs = [item['전표번호'] for item in invoice_matched]
            for _, row in invoice_docs.iterrows():
                if row['전표번호'] not in matched_docs and row['전표번호'] not in grouped['Doc.No.(중복제거)'].values:
                    invoice_unmatched.append({
                        '전표번호': row['전표번호'],
                        '금액': row['금액']
                    })
        
        df_invoice = pd.DataFrame(invoice_matched + invoice_unmatched)
        if df_invoice.empty:
            df_invoice = pd.DataFrame(columns=['전표번호', '금액'])
        
        # 최종 DataFrame 조합
        max_rows = max(len(df_col1), len(grouped), len(df_invoice))
        
        result = pd.DataFrame()
        result['Doc.No.'] = df_col1['Doc.No.'].reindex(range(max_rows))
        result['AI-AJ'] = df_col1['Amount'].reindex(range(max_rows))
        result['Doc.Date'] = df_col1['Doc.Date'].reindex(range(max_rows))
        result[''] = None
        result['Doc.No.(중복제거)'] = grouped['Doc.No.(중복제거)'].reindex(range(max_rows))
        result['합계'] = grouped['합계'].reindex(range(max_rows))
        result['Doc.Date.1'] = grouped['Doc.Date'].reindex(range(max_rows))
        result[' '] = None
        result['전표번호'] = df_invoice['전표번호'].reindex(range(max_rows))
        result['금액'] = df_invoice['금액'].reindex(range(max_rows))
        
        return result
    
    def _create_final_report_headers(self, ws):
        """최종 보고서 헤더 생성"""
        ws.merge_cells('A1:S1')
        ws['A1'] = '외화획득내용'
        
        ws.merge_cells('A2:C2')
        ws['A2'] = '공급일자'
        ws.merge_cells('D2:G2')
        ws['D2'] = '공급받는 자'
        ws.merge_cells('H2:N2')
        ws['H2'] = '공급내용'
        
        ws.merge_cells('O2:O5')
        ws['O2'] = '비고'
        ws.merge_cells('P2:P5')
        ws['P2'] = '전표번호'
        ws.merge_cells('Q2:Q5')
        ws['Q2'] = '외화'
        ws.merge_cells('R2:R5')
        ws['R2'] = '단위'
        ws.merge_cells('S2:S5')
        ws['S2'] = '순번'
        
        self._create_detailed_headers(ws)
    
    def _create_detailed_headers(self, ws):
        """세부 헤더 생성"""
        ws.merge_cells('A3:A5')
        ws['A3'] = '년'
        ws.merge_cells('B3:B5')
        ws['B3'] = '월'
        ws.merge_cells('C3:C5')
        ws['C3'] = '일'
        
        ws.merge_cells('D3:F5')
        ws['D3'] = '상호 및 성명'
        ws.merge_cells('G3:G5')
        ws['G3'] = '국적'
        
        ws.merge_cells('H3:I3')
        ws['H3'] = '구분'
        ws.merge_cells('H4:H5')
        ws['H4'] = '재화'
        ws.merge_cells('I4:I5')
        ws['I4'] = '용역'
        
        ws.merge_cells('J3:J5')
        ws['J3'] = '명칭'
        ws.merge_cells('K3:K5')
        ws['K3'] = '수량'
        ws.merge_cells('L3:M5')
        ws['L3'] = '단가'
        ws.merge_cells('N3:N5')
        ws['N3'] = '공급가액'
        
        ws.merge_cells('T4:T5')
        ws['T4'] = 'ERP 환율확인'
        ws.merge_cells('U4:U5')
        ws['U4'] = '매매기준 환율'
        ws.merge_cells('V4:V5')
        ws['V4'] = '환율 차이확인'
        ws.merge_cells('W4:W5')
        ws['W4'] = 'KRW x 환율'
        ws.merge_cells('X4:X5')
        ws['X4'] = 'A2대사'
        
        ws['Y5'] = '인보이스 번호'
        ws['Z5'] = '고객사명'
        ws['AA5'] = '날짜'
        ws['AB5'] = 'KRW계약여부'
        ws['AC5'] = 'A2원화'
        
        ws['AB2'] = '합계'
    
    def _fill_final_report_data(self, ws, invoice_data: pd.DataFrame, 
                               exchange_data: Optional[pd.DataFrame], 
                               a2_data: Optional[pd.DataFrame]):
        """최종 보고서 데이터 입력"""
        if invoice_data is None or invoice_data.empty:
            return
        
        start_row = 6
        
        for idx, row in invoice_data.iterrows():
            current_row = start_row + idx
            
            # 공급일자
            if pd.notna(row.get('Doc. Date')):
                date_val = pd.to_datetime(row['Doc. Date'])
                ws.cell(row=current_row, column=1, value=date_val.year)
                ws.cell(row=current_row, column=2, value=date_val.month)
                ws.cell(row=current_row, column=3, value=date_val.day)
            
            # 상호 및 성명
            ws.merge_cells(f'D{current_row}:F{current_row}')
            ws.cell(row=current_row, column=4, value=row.get('Customer', ''))
            
            # 국적
            ws.cell(row=current_row, column=7, value=row.get('국가', ''))
            
            # 구분 표시
            if row.get('구분') == '재화':
                ws.cell(row=current_row, column=8, value='●')
                ws.cell(row=current_row, column=10, value='국외제공재화')
            elif row.get('구분') == '용역':
                ws.cell(row=current_row, column=9, value='●')
                ws.cell(row=current_row, column=10, value='국외제공용역')
            
            # 비고, 전표번호, 외화, 단위, 순번
            ws.cell(row=current_row, column=15, value='단위:원')
            ws.cell(row=current_row, column=16, value=row.get('전표번호', ''))
            amount = row.get('금액', 0)
            ws.cell(row=current_row, column=17, value=amount)
            currency = row.get('단위', '')
            ws.cell(row=current_row, column=18, value=currency)
            ws.cell(row=current_row, column=19, value=idx + 1)
            
            # 매매기준 환율
            exchange_rate = None
            if currency == 'KRW':
                exchange_rate = 1
                ws.cell(row=current_row, column=21, value=1)
            elif exchange_data is not None and pd.notna(row.get('Doc. Date')) and currency:
                exchange_rate = self._get_exchange_rate(exchange_data, currency, row['Doc. Date'])
                if exchange_rate:
                    ws.cell(row=current_row, column=21, value=exchange_rate)
            
            # 공급가액
            if exchange_rate and amount:
                supply_value = amount * exchange_rate
                ws.cell(row=current_row, column=14, value=supply_value)
            
            # 추가 정보
            ws.cell(row=current_row, column=25, value=row.get('인보이스 넘버', ''))
            ws.cell(row=current_row, column=26, value=row.get('Customer', ''))
            if pd.notna(row.get('Doc. Date')):
                ws.cell(row=current_row, column=27, value=row['Doc. Date'])
            
            # A2원화
            if a2_data is not None:
                doc_no = row.get('전표번호', '')
                if doc_no:
                    a2_amount = self._get_a2_amount(a2_data, doc_no)
                    if a2_amount is not None:
                        ws.cell(row=current_row, column=29, value=a2_amount)
            
            self._calculate_additional_columns(ws, current_row)
    
    def _get_exchange_rate(self, exchange_data: pd.DataFrame, currency: str, date: datetime) -> Optional[float]:
        """특정 날짜의 환율 조회"""
        if currency not in exchange_data.columns:
            return None
        
        try:
            date_only = pd.to_datetime(date).date()
            matching_rows = exchange_data[exchange_data['일자'].dt.date == date_only]
            
            if not matching_rows.empty:
                return matching_rows.iloc[0][currency]
            return None
        except:
            return None
    
    def _get_a2_amount(self, a2_data: pd.DataFrame, doc_no: str) -> Optional[float]:
        """A2 데이터에서 금액 조회"""
        try:
            doc_no_str = str(doc_no).strip()
            if doc_no_str.endswith('.0'):
                doc_no_str = doc_no_str[:-2]
            
            if 'Doc.No.(중복제거)' in a2_data.columns:
                a2_data_copy = a2_data.copy()
                a2_data_copy['Doc.No.(중복제거)'] = (
                    a2_data_copy['Doc.No.(중복제거)']
                    .astype(str)
                    .str.strip()
                    .str.replace('.0', '')
                    .str.replace('.00', '')
                )
                
                matching_rows = a2_data_copy[a2_data_copy['Doc.No.(중복제거)'] == doc_no_str]
                
                if not matching_rows.empty and '합계' in a2_data.columns:
                    return matching_rows.iloc[0]['합계']
            
            return None
        except:
            return None
    
    def _calculate_additional_columns(self, ws, current_row):
        """추가 계산 컬럼들 처리"""
        ac_val = ws.cell(row=current_row, column=29).value
        q_val = ws.cell(row=current_row, column=17).value
        if ac_val and q_val and q_val != 0:
            ws.cell(row=current_row, column=20, value=ac_val / q_val)
        
        t_val = ws.cell(row=current_row, column=20).value
        u_val = ws.cell(row=current_row, column=21).value
        if t_val is not None and u_val is not None:
            ws.cell(row=current_row, column=22, value=t_val - u_val)
        
        n_val = ws.cell(row=current_row, column=14).value
        if u_val and q_val:
            calc_val = u_val * q_val
            if n_val and abs(calc_val - n_val) < 0.01:
                ws.cell(row=current_row, column=23, value='TRUE')
            else:
                ws.cell(row=current_row, column=23, value='FALSE')
        
        if ac_val is not None and n_val is not None:
            diff_value = ac_val - n_val
            
            if diff_value == 0:
                display_value = "-"
            elif abs(diff_value) < 1:
                display_value = "0"
            else:
                display_value = diff_value
            
            cell = ws.cell(row=current_row, column=24, value=display_value)
            
            if diff_value != 0:
                from openpyxl.styles import Font
                cell.font = Font(color="FF0000")
    
    def _apply_final_report_styles(self, ws):
        """최종 보고서 스타일 적용"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        header_font = Font(size=14, bold=True, color="FFFFFF")
        sub_header_font = Font(size=11, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        # 헤더 스타일
        for row in range(1, 6):
            for col in range(1, 30):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    if row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        cell.fill = sub_header_fill
                        cell.font = sub_header_font
                    cell.alignment = center_align
                    cell.border = thin_border
        
        # 데이터 영역 스타일
        last_row = ws.max_row
        for row in range(6, last_row + 1):
            for col in range(1, 30):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                if col in [14, 17, 20, 21, 22, 29]:
                    cell.number_format = '#,##0.00'
                elif col == 27:
                    cell.number_format = 'YYYY-MM-DD'
    
    def _add_final_report_formulas(self, ws):
        """최종 보고서 수식 추가"""
        last_row = ws.max_row
        
        if last_row >= 6:
            ws['AC2'] = f'=SUM(N6:N{last_row})'
            ws['AC2'].font = Font(bold=True, size=12)
            ws['AC2'].number_format = '#,##0'
    
    def _create_excel_with_logs(self, data_df: pd.DataFrame, logs: List[Dict], 
                               data_sheet_name: str = '데이터', 
                               log_sheet_name: str = '로그') -> bytes:
        """데이터와 로그를 포함한 Excel 파일 생성"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data_df.to_excel(writer, sheet_name=data_sheet_name, index=False)
            
            if logs:
                log_df = pd.DataFrame(logs)
                log_df.to_excel(writer, sheet_name=log_sheet_name, index=False)
        
        return output.getvalue()
    
    def _create_a2_excel(self, data_df: pd.DataFrame) -> bytes:
        """A2 정리 전용 Excel 파일 생성"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data_df.to_excel(writer, sheet_name='A2(정리)', index=False)
        
        return output.getvalue()