"""
TAXi - 원천세 데이터 처리기
기존 원천세 시스템의 데이터 처리 로직을 TAXi 구조에 맞게 수정
"""

import pandas as pd
import logging
import os
from datetime import datetime
import math
import streamlit as st
import io
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config.withholding_tax_config import COMPANY_CONFIGS, WORKPLACE_CONFIGS, WORKPLACE_ORDER, INCOME_TYPES, HOMETAX_ITEMS
from utils.withholding_tax.web_helpers import create_empty_data, safe_float, format_number, find_row_by_keywords, aggregate_data
import logging

class WithholdingTaxProcessor:
    """원천세 처리 메인 클래스"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.company4_processor = Company4Processor()
        self.excel_generator = ExcelGenerator()

    def process(self, uploaded_file, company_type=None, sheet_name=None):
        """업로드된 파일 처리"""
        try:
            if company_type is None:
                company_type = self.detect_company_type(uploaded_file.name)
                if company_type is None:
                    company_type = 1  # 기본값으로 애커튼 설정
                    self.logger.warning(f"회사 타입을 인식할 수 없어 기본값(애커튼)으로 설정: {uploaded_file.name}")
    

            temp_path = self._save_to_temp_file(uploaded_file)
            if not temp_path:
                return {
                    'success': False,
                    'error': "임시 파일 생성 실패",
                    'data': {},
                    'excel_data': b''
                }

            try:
                data_summary, errors = self.process_company_data(temp_path, company_type, sheet_name)
                excel_data = self.excel_generator.create_excel_bytes(data_summary, company_type)
                
                return {
                    'success': True,
                    'error': None,
                    'data': data_summary,
                    'excel_data': excel_data,
                    'errors': errors,
                    'company_type': company_type,
                    'company_name': COMPANY_CONFIGS[company_type]['name']
                }
            finally:
                self._cleanup_temp_file(temp_path)

        except Exception as e:
            error_msg = f"파일 처리 중 오류: {str(e)}"
            self.logger.error(error_msg)
            return {
                'success': False,
                'error': error_msg,
                'data': {},
                'excel_data': b''
            }

    def process_combined(self, file_company_mappings):
        """통합 파일 처리"""
        try:
            all_companies_data = {}
            all_errors = []

            for uploaded_file, company_type, sheet_name in file_company_mappings:
                result = self.process(uploaded_file, company_type, sheet_name)
                
                if result['success']:
                    all_companies_data[company_type] = result['data']
                    if 'errors' in result and result['errors']:
                        all_errors.extend(result['errors'])
                else:
                    all_errors.append(f"{COMPANY_CONFIGS[company_type]['name']}: {result['error']}")

            combined_excel_data = self.excel_generator.create_combined_excel_bytes(all_companies_data)

            return {
                'success': True,
                'error': None,
                'data': all_companies_data,
                'excel_data': combined_excel_data,
                'errors': all_errors
            }

        except Exception as e:
            error_msg = f"통합 처리 중 오류: {str(e)}"
            self.logger.error(error_msg)
            return {
                'success': False,
                'error': error_msg,
                'data': {},
                'excel_data': b''
            }

    def process_company_data(self, file_path, company_type, sheet_name=None):
        """회사별 데이터 처리"""
        config = COMPANY_CONFIGS[company_type]
        errors = []

        try:
            if company_type == 4:
                return self.company4_processor.process_company4(file_path, config, errors)
            else:
                return self._process_standard_company(file_path, company_type, config, sheet_name, errors)
        except Exception as e:
            error_msg = f"파일 처리 중 오류: {str(e)}"
            errors.append(error_msg)
            self.logger.error(error_msg)
            return {}, errors
        
            # 여기에 추가
            data_summary = {}
            if not data_summary:
                self.logger.warning("추출된 데이터가 없습니다. 기본 구조를 생성합니다.")
                data_summary = {'총합계': create_empty_data()}
                
            return data_summary, errors

    def _process_standard_company(self, file_path, company_type, config, sheet_name, errors):
        """표준 회사(1,2,3번) 데이터 처리"""
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        data_summary = {}

        extractor = DataExtractor()
        extract_method = getattr(extractor, f"extract_company{company_type}")
        
        for key, row_idx, prefix in config['mappings']:
            data_summary[key] = extract_method(df, row_idx, prefix)

        self._apply_special_processing(company_type, config, data_summary)
        self._calculate_totals(company_type, config, data_summary)

        # 여기에 추가
        if not data_summary:
            self.logger.warning("추출된 데이터가 없습니다. 기본 구조를 생성합니다.")
            data_summary = {'총합계': create_empty_data()}

        return data_summary, errors


    def _apply_special_processing(self, company_type, config, data_summary):
        """특수 처리 적용"""
        if company_type == 3 and 'zero_items' in config:
            for item in config['zero_items']:
                data_summary[item] = create_empty_data()

        if company_type == 1:
            data_summary['퇴직소득 과세'] = create_empty_data()

    def _calculate_totals(self, company_type, config, data_summary):
        """합계 계산"""
        groups = config['groups']
        
        for group_name, keys in groups.items():
            if group_name == 'other':
                continue

            total = create_empty_data()
            for key in keys:
                if key in data_summary:
                    for field in total:
                        total[field] += data_summary[key][field]

            total_name = self._get_total_name(company_type, group_name)
            if total_name:
                data_summary[total_name] = total

            if company_type in [2, 3] and group_name in ['salary', 'mid_resignation']:
                if '근로소득 총합계' not in data_summary:
                    data_summary['근로소득 총합계'] = create_empty_data()
                for field in total:
                    data_summary['근로소득 총합계'][field] += total[field]

        self._calculate_grand_total(company_type, data_summary)

    def _get_total_name(self, company_type, group_name):
        """그룹별 합계명 반환"""
        mapping = {
            1: {
                'salary': '급여 총합계',
                'retirement': '퇴직소득 총합계',
                'yearend': '연말정산 전체 합계'
            },
            2: {
                'salary': '급여 총합계',
                'mid_resignation': '중도퇴사연말정산 총합계',
                'yearend': '연말정산 총합계',
                'retirement': '퇴직소득 총합계'
            },
            3: {
                'salary': '급여 총합계',
                'mid_resignation': '중도퇴사연말정산 총합계',
                'yearend': '연말정산 총합계',
                'retirement': '퇴직소득 총합계'
            }
        }
        return mapping.get(company_type, {}).get(group_name)

    def _calculate_grand_total(self, company_type, data_summary):
        """전체 합계 계산"""
        total = create_empty_data()

        for key, data in data_summary.items():
            if not any(keyword in key for keyword in ['총합계', '합계']):
                for field in total:
                    total[field] += data[field]

        data_summary['총합계'] = total

        if company_type == 3:
            final_total = create_empty_data()
            for field in final_total:
                final_total[field] = total[field]
                if '연말정산 총합계' in data_summary:
                    final_total[field] += data_summary['연말정산 총합계'][field]
            data_summary['최종 총합계'] = final_total

    def detect_company_type(self, filename):
        """파일명 기반 회사 타입 감지"""
        filename_upper = filename.upper()

        if "AX" in filename_upper:
            return 4
        elif "MR" in filename_upper or "MRCIC" in filename_upper:
            return 3
        elif "INC" in filename_upper:
            return 2
        elif "애커튼" in filename or "ACCERTON" in filename_upper:
            return 1
        else:
            return None

    def _save_to_temp_file(self, uploaded_file):
        """업로드된 파일을 임시 파일로 저장"""
        try:
            # TAXi 구조에 맞게 data/temp 사용
            temp_dir = os.path.join(os.getcwd(), 'data', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_id = st.session_state.get('wt_session_id', 'default')
            temp_filename = f"temp_{session_id}_{timestamp}_{uploaded_file.name}"
            temp_path = os.path.join(temp_dir, temp_filename)
            
            with open(temp_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())
            
            return temp_path
        except Exception as e:
            self.logger.error(f"임시 파일 저장 중 오류: {e}")
            return None

    def _cleanup_temp_file(self, temp_path):
        """임시 파일 정리"""
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            self.logger.error(f"임시 파일 정리 중 오류: {e}")


class DataExtractor:
    """데이터 추출 클래스"""
    
    def extract_company1(self, df, row_idx, prefix):
        """1번째 회사 (애커튼 파트너스) 데이터 추출"""
        try:
            columns = {
                '인원': 5, '소득세': 9, '주민세': 10,
                '과세소득_alt': 6, '과세소득': 13, '제출비과세': 14
            }
            data = self._extract_base_data(df, row_idx, columns)
            
            if data['과세소득'] and data['제출비과세']:
                taxable = data['과세소득']
                nontaxable = data['제출비과세']
            else:
                taxable = data['과세소득_alt']
                nontaxable = 0
            
            return {
                '인원': data['인원'],
                '과세소득': taxable,
                '제출비과세': nontaxable,
                '총지급액': taxable + nontaxable,
                '소득세': data['소득세'],
                '주민세': data['주민세']
            }
        except Exception as e:
            logging.warning(f"{prefix} (행 {row_idx+1}) 데이터 형식 오류: {str(e)}")
            return create_empty_data()

    def extract_company2(self, df, row_idx, prefix):
        """2번째 회사 (SK INC) 데이터 추출"""
        try:
            columns = {'인원': 5, '과세소득': 6, '소득세': 7, '주민세': 8}
            data = self._extract_base_data(df, row_idx, columns)
            
            return {
                '인원': data['인원'],
                '과세소득': data['과세소득'],
                '제출비과세': 0,
                '총지급액': data['과세소득'],
                '소득세': data['소득세'],
                '주민세': data['주민세']
            }
        except Exception as e:
            logging.warning(f"{prefix} (행 {row_idx+1}) 데이터 형식 오류: {str(e)}")
            return create_empty_data()

    def extract_company3(self, df, row_idx, prefix):
        """3번째 회사 (SK MRCIC) 데이터 추출"""
        try:
            columns = {
                '인원': 5, '소득세': 9, '주민세': 10,
                '과세소득': 6, '제출비과세': 7
            }
            data = self._extract_base_data(df, row_idx, columns)
            
            return {
                '인원': data['인원'],
                '과세소득': data['과세소득'],
                '제출비과세': data['제출비과세'],
                '총지급액': data['과세소득'] + data['제출비과세'],
                '소득세': data['소득세'],
                '주민세': data['주민세']
            }
        except Exception as e:
            logging.warning(f"{prefix} (행 {row_idx+1}) 데이터 형식 오류: {str(e)}")
            return create_empty_data()

    def _extract_base_data(self, df, row_idx, columns):
        """기본 데이터 추출 공통 함수"""
        if row_idx >= len(df):
            return create_empty_data()
        
        row_data = df.iloc[row_idx:row_idx+1]
        if row_data.empty:
            return create_empty_data()
        
        return {key: safe_float(row_data.iloc[0, col]) for key, col in columns.items()}


class Company4Processor:
    """SK AX (4번 회사) 전용 데이터 처리 모듈"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def process_company4(self, file_path, config, errors):
        """4번째 회사 특수 처리"""
        excel_file = pd.ExcelFile(file_path)
        available_sheets = excel_file.sheet_names
        data_summary = {}

        personnel_data = self._extract_personnel_data(file_path, available_sheets, errors)

        sheet_processors = {
            '원천세신고내역(표)': self._process_salary_sheet,
            '아르바이트': self._process_parttime_sheet,
            '근로소득': self._process_employment_sheet,
            '세액': self._process_tax_sheet,
            '퇴직소득': self._process_retirement_sheet,
            '기타&사업소득원장': self._process_other_income_sheet,
            '배당소득': self._process_dividend_sheet,
            '연말정산': self._process_yearend_sheet
        }

        for sheet_name, items in config['sheets'].items():
            if sheet_name in available_sheets and sheet_name in sheet_processors:
                try:
                    if sheet_name == '원천세신고내역(표)':
                        sheet_processors[sheet_name](file_path, sheet_name, items, data_summary, personnel_data, errors)
                    else:
                        sheet_processors[sheet_name](file_path, items, data_summary, errors)
                except Exception as e:
                    error_msg = f"시트 {sheet_name} 처리 실패: {str(e)}"
                    errors.append(error_msg)
                    self.logger.error(error_msg)

        self._calculate_company4_totals(config, data_summary)
        return data_summary, errors
    
    def _extract_personnel_data(self, file_path, available_sheets, errors):
        """인원수 데이터 추출"""
        personnel_data = {}
        if '인원수' in available_sheets:
            try:
                df_personnel = pd.read_excel(file_path, sheet_name='인원수', header=None)
                
                region_keywords = {
                    '분당': ['분당'],
                    '충무로': ['충무로', '종로'],
                    '대덕': ['대덕']
                }
                
                for region, keywords in region_keywords.items():
                    row_idx = find_row_by_keywords(df_personnel, keywords, search_columns=[0])
                    if row_idx is not None:
                        personnel_data[region] = int(df_personnel.iloc[row_idx, 1]) if pd.notna(df_personnel.iloc[row_idx, 1]) else 0
                    else:
                        personnel_data[region] = 0
                        
            except Exception as e:
                errors.append(f"인원수 시트 처리 실패: {str(e)}")
        return personnel_data

    # ... (나머지 Company4Processor 메서드들은 기존과 동일) ...
    
    def _process_salary_sheet(self, file_path, sheet_name, items, data_summary, personnel_data, errors):
        """급여 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            df_other = None
            try:
                df_other = pd.read_excel(file_path, sheet_name='기타근로', header=None)
            except:
                pass

            # SK AX 전용 지역 매핑 (충무로/종로 대안 키워드)
            ax_region_mapping = {
                '급여(분당)': ['분당'],
                '급여(충무로)': ['충무로', '종로'],  # '종로', '종로 오피스', '종로오피스' 모두 매치
                '급여(대덕)': ['대덕']
            }
            
            region_map = {'급여(분당)': '분당', '급여(충무로)': '충무로', '급여(대덕)': '대덕'}

            for item in items:
                keywords = ax_region_mapping.get(item, [item])
                row_idx = find_row_by_keywords(df, keywords, search_columns=[0, 1, 2])
                
                if row_idx is not None:
                    taxable_income = safe_float(df.iloc[row_idx, 3])
                    nontaxable_income = safe_float(df.iloc[row_idx, 6])
                    income_tax = safe_float(df.iloc[row_idx, 8])
                    resident_tax = safe_float(df.iloc[row_idx, 9])

                    # 급여(분당)의 경우 기타근로 합산
                    if item == '급여(분당)' and df_other is not None:
                        try:
                            h_column = df_other.iloc[2:, 7]
                            numeric_values = h_column.dropna()
                            other_sum = 0
                            for val in numeric_values:
                                try:
                                    other_sum += float(val)
                                except (ValueError, TypeError):
                                    continue
                            taxable_income += other_sum
                        except:
                            pass

                    region = region_map.get(item, '')
                    personnel = personnel_data.get(region, 0)

                    data_summary[item] = {
                        '인원': personnel,
                        '과세소득': taxable_income,
                        '제출비과세': nontaxable_income,
                        '총지급액': taxable_income + nontaxable_income,
                        '소득세': income_tax,
                        '주민세': resident_tax
                    }
                else:
                    data_summary[item] = create_empty_data()
        except Exception as e:
            error_msg = f"급여 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_parttime_sheet(self, file_path, items, data_summary, errors):
        """아르바이트 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='아르바이트', header=None)
            header_row = df.iloc[1] if len(df) > 1 else pd.Series()

            # 컬럼 찾기
            col_mapping = {}
            for col_name in ['총지급액', '소득세', '지방소득세']:
                try:
                    col_mapping[col_name] = header_row[header_row == col_name].index[0]
                except IndexError:
                    col_mapping[col_name] = None

            for item in items:
                pay_type = '월급' if item == '일용직(3개월이상)' else '일급'
                mask = df[2].str.contains(pay_type, na=False) if 2 in df.columns else pd.Series([False] * len(df))

                if mask.any():
                    unique_ids = df[mask][4].drop_duplicates() if 4 in df.columns else pd.Series()
                    personnel_count = len(unique_ids)
                    
                    total_payment = df[mask][col_mapping['총지급액']].sum() if col_mapping['총지급액'] is not None else 0
                    income_tax = df[mask][col_mapping['소득세']].sum() if col_mapping['소득세'] is not None else 0
                    local_tax = df[mask][col_mapping['지방소득세']].sum() if col_mapping['지방소득세'] is not None else 0

                    data_summary[item] = {
                        '인원': personnel_count,
                        '과세소득': safe_float(total_payment),
                        '제출비과세': 0,
                        '총지급액': safe_float(total_payment),
                        '소득세': safe_float(income_tax),
                        '주민세': safe_float(local_tax)
                    }
                else:
                    data_summary[item] = create_empty_data()
        except Exception as e:
            error_msg = f"아르바이트 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_employment_sheet(self, file_path, items, data_summary, errors):
        """근로소득 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='근로소득', header=None)
            exclude_keywords = ['임원', '관계사', '전출', '승계', 'SK', '사망']
            retirement_keyword = '정년'

            # 마스크 생성
            mask_exclude = pd.Series(False, index=df.index)
            mask_retirement = pd.Series(False, index=df.index)

            for col in range(min(9, len(df.columns))):
                if col in df.columns:
                    mask_exclude |= df[col].str.contains('|'.join(exclude_keywords), na=False)
                    mask_retirement |= df[col].str.contains(retirement_keyword, na=False)

            # 인원 계산
            personnel_mask = ~mask_exclude & ~mask_retirement
            personnel_data_mid = df[0][1:].dropna()[personnel_mask[1:]].drop_duplicates() if 0 in df.columns else pd.Series()
            personnel_count = len(personnel_data_mid)

            # 금액 계산
            amount_mask = ~mask_exclude
            taxable_income = df[3][1:].dropna()[amount_mask[1:]].astype(float).sum() if 3 in df.columns else 0
            income_tax = df[5][1:].dropna()[amount_mask[1:]].astype(float).sum() if 5 in df.columns else 0
            local_tax = df[6][1:].dropna()[amount_mask[1:]].astype(float).sum() if 6 in df.columns else 0

            data_summary['중도퇴사연말정산(분당)'] = {
                '인원': personnel_count,
                '과세소득': taxable_income,
                '제출비과세': 0,
                '총지급액': taxable_income,
                '소득세': income_tax,
                '주민세': local_tax
            }
        except Exception as e:
            error_msg = f"근로소득 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_tax_sheet(self, file_path, items, data_summary, errors):
        """세액 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='세액', header=None)

            income_tax_series = df[8][2:].dropna().astype(float) if 8 in df.columns else pd.Series()
            income_tax_series = income_tax_series[income_tax_series != 0]
            personnel_count = len(income_tax_series)
            income_tax = income_tax_series.sum()

            local_tax_series = df[9][2:].dropna().astype(float) if 9 in df.columns else pd.Series()
            local_tax_series = local_tax_series[local_tax_series != 0]
            local_tax = local_tax_series.sum()

            if income_tax != 0 or local_tax != 0:
                errors.append("우리사주 확인필요: 소득세 또는 지방소득세가 0이 아님")

            data_summary['우리사주(분당)'] = {
                '인원': personnel_count,
                '과세소득': 0,
                '제출비과세': 0,
                '총지급액': 0,
                '소득세': income_tax,
                '주민세': local_tax
            }
        except Exception as e:
            error_msg = f"세액 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_retirement_sheet(self, file_path, items, data_summary, errors):
        """퇴직소득 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='퇴직소득', header=None)

            # 과세: E열 0이 아닌 행
            mask_taxed = df[4][1:].dropna().astype(float) != 0 if 4 in df.columns else pd.Series([False] * (len(df)-1))
            taxed_personnel = df[0][1:].dropna()[mask_taxed].drop_duplicates() if 0 in df.columns else pd.Series()
            taxed_personnel_count = len(taxed_personnel)
            taxed_taxable_income = df[3][1:].dropna()[mask_taxed].astype(float).sum() if 3 in df.columns else 0
            taxed_income_tax = df[4][1:].dropna()[mask_taxed].astype(float).sum() if 4 in df.columns else 0
            taxed_local_tax = df[5][1:].dropna()[mask_taxed].astype(float).sum() if 5 in df.columns else 0

            data_summary['퇴직소득 과세'] = {
                '인원': taxed_personnel_count,
                '과세소득': taxed_taxable_income,
                '제출비과세': 0,
                '총지급액': taxed_taxable_income,
                '소득세': taxed_income_tax,
                '주민세': taxed_local_tax
            }

            # 과세이연: E열 0이면서 D열이 0이 아닌 행
            mask_deferred = ((df[4][1:].dropna().astype(float) == 0) & 
                           (df[3][1:].dropna().astype(float) != 0)) if 3 in df.columns and 4 in df.columns else pd.Series([False] * (len(df)-1))
            deferred_personnel = df[0][1:].dropna()[mask_deferred].drop_duplicates() if 0 in df.columns else pd.Series()
            deferred_personnel_count = len(deferred_personnel)
            deferred_taxable_income = df[3][1:].dropna()[mask_deferred].astype(float).sum() if 3 in df.columns else 0

            data_summary['퇴직소득 과세이연'] = {
                '인원': deferred_personnel_count,
                '과세소득': deferred_taxable_income,
                '제출비과세': 0,
                '총지급액': deferred_taxable_income,
                '소득세': 0,
                '주민세': 0
            }
        except Exception as e:
            error_msg = f"퇴직소득 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_other_income_sheet(self, file_path, items, data_summary, errors):
        """기타&사업소득원장 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='기타&사업소득원장', header=None)

            for item in items:
                income_type = item
                mask = df[1].str.contains(income_type, na=False) if 1 in df.columns else pd.Series([False] * len(df))
                personnel_count = len(df[mask])
                taxable_income = df[4][mask].dropna().astype(float).sum() if 4 in df.columns else 0
                income_tax = df[5][mask].dropna().astype(float).sum() if 5 in df.columns else 0
                local_tax = df[6][mask].dropna().astype(float).sum() if 6 in df.columns else 0

                data_summary[item] = {
                    '인원': personnel_count,
                    '과세소득': taxable_income,
                    '제출비과세': 0,
                    '총지급액': taxable_income,
                    '소득세': income_tax,
                    '주민세': local_tax
                }
        except Exception as e:
            error_msg = f"기타&사업소득원장 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_dividend_sheet(self, file_path, items, data_summary, errors):
        """배당소득 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='배당소득', header=None)
            taxable_income = df[5][2:].dropna().astype(float).sum() if 5 in df.columns else 0

            data_summary['배당소득'] = {
                '인원': 0,
                '과세소득': taxable_income,
                '제출비과세': 0,
                '총지급액': taxable_income,
                '소득세': 0,
                '주민세': 0
            }
        except Exception as e:
            error_msg = f"배당소득 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _process_yearend_sheet(self, file_path, items, data_summary, errors):
        """연말정산 시트 처리"""
        try:
            df = pd.read_excel(file_path, sheet_name='연말정산', header=None)

            for item in items:
                region = item.replace('연말정산(', '').replace(')', '')
                
                # SK AX 전용: 충무로는 종로로도 검색
                if region == '충무로':
                    keywords = ['충무로', '종로']
                else:
                    keywords = [region]
                
                row_idx = find_row_by_keywords(df, keywords, search_columns=[0, 1, 2])
                
                if row_idx is not None:
                    personnel_count = safe_float(df.iloc[row_idx, 5]) if 5 in df.columns else 0
                    taxable_income = safe_float(df.iloc[row_idx, 2]) if 2 in df.columns else 0
                    income_tax = safe_float(df.iloc[row_idx, 3]) if 3 in df.columns else 0
                    local_tax = safe_float(df.iloc[row_idx, 4]) if 4 in df.columns else 0

                    data_summary[item] = {
                        '인원': personnel_count,
                        '과세소득': taxable_income,
                        '제출비과세': 0,
                        '총지급액': taxable_income,
                        '소득세': income_tax,
                        '주민세': local_tax
                    }
                else:
                    data_summary[item] = create_empty_data()
        except Exception as e:
            error_msg = f"연말정산 시트 처리 실패: {str(e)}"
            errors.append(error_msg)
            for item in items:
                data_summary[item] = create_empty_data()

    def _calculate_company4_totals(self, config, data_summary):
        """4번째 회사 합계 계산"""
        groups = config['groups']
        group_total_names = {
            'salary': '급여 총합계',
            'mid_resignation': '중도퇴사연말정산 총합계',
            'employee_stock': '우리사주 총합계',
            'retirement': '퇴직소득 총합계',
            'yearend': '연말정산 합계'
        }

        for group_name, keys in groups.items():
            if group_name == 'other':
                continue
                
            total = create_empty_data()
            for key in keys:
                if key in data_summary:
                    for field in total:
                        # 우리사주는 과세소득, 총지급액 제외
                        if field in ['과세소득', '총지급액'] and key == '우리사주(분당)':
                            continue
                        total[field] += data_summary[key][field]

            if group_name in group_total_names:
                data_summary[group_total_names[group_name]] = total

        # 근로소득 총합계
        data_summary['근로소득 총합계'] = create_empty_data()
        for group in ['salary', 'mid_resignation', 'employee_stock']:
            group_key = group_total_names[group]
            if group_key in data_summary:
                for field in data_summary['근로소득 총합계']:
                    data_summary['근로소득 총합계'][field] += data_summary[group_key][field]

        # 총합계
        total = create_empty_data()
        for key in ['근로소득 총합계', '퇴직소득 총합계'] + [k for k in data_summary.keys() if k in groups['other']]:
            if key in data_summary:
                for field in total:
                    total[field] += data_summary[key][field]
        data_summary['총합계'] = total

        # 최종 총합계 (연말정산 포함)
        final_total = create_empty_data()
        for field in final_total:
            final_total[field] = total[field]
            if '연말정산 합계' in data_summary:
                final_total[field] += data_summary['연말정산 합계'][field]
        data_summary['최종 총합계'] = final_total


class ExcelGenerator:
    """엑셀 파일 생성 클래스"""

    def _aggregate_workplace_data(self, all_companies_data):
        """사업장별 데이터 집계"""
        workplace_summary = {}
        
        # 각 사업장별로 빈 데이터 구조 초기화
        for workplace in WORKPLACE_ORDER:
            workplace_summary[workplace] = {}
            for income_type in INCOME_TYPES:
                workplace_summary[workplace][income_type] = {'인원': 0, '소득세': 0, '주민세': 0}
        
        # 회사별 데이터를 사업장별로 집계
        for company_type, company_data in all_companies_data.items():
            if company_type == 1:  # 애커튼
                workplace_name = '애커튼'
                self._aggregate_company_to_workplace(workplace_summary[workplace_name], company_data, company_type)
            elif company_type == 2:  # SK INC
                # INC본점
                inc_items = ['급여(INC 본점)', '일용근로소득', '중도퇴사연말정산(INC 본점)', '퇴직소득 과세', '기타소득', '사업소득', '법인원천소득', '배당소득', '연말정산납부금액']
                self._aggregate_specific_items(workplace_summary['INC본점'], company_data, inc_items)
                
                # 이천
                icheon_items = ['급여(이천)', '중도퇴사연말정산(이천)', '연말정산이천']
                self._aggregate_specific_items(workplace_summary['이천'], company_data, icheon_items)
                
            elif company_type == 3:  # SK MRCIC
                # 각 사업장별 항목 집계
                mr_mappings = {
                    'MR본점': ['급여(MR 본점)', '일용근로소득', '중도퇴사연말정산(MR 본점)', '퇴직소득 과세', '기타소득', '사업소득', '법인원천소득', '배당소득', '연말정산(MR 본점)'],
                    '세종': ['급여(세종)', '중도퇴사연말정산(세종)', '연말정산(세종)'],
                    '영주': ['급여(영주)', '중도퇴사연말정산(영주)', '연말정산(영주)'],
                    '동탄': ['급여(동탄)', '중도퇴사연말정산(동탄)', '연말정산(동탄)'],
                    '상주': ['급여(상주)', '중도퇴사연말정산(상주)', '연말정산(상주)']
                }
                for workplace, items in mr_mappings.items():
                    self._aggregate_specific_items(workplace_summary[workplace], company_data, items)
                    
            elif company_type == 4:  # SK AX
                # 각 사업장별 항목 집계
                ax_mappings = {
                    '분당': ['급여(분당)', '일용직(3개월이상)', '일용근로소득', '중도퇴사연말정산(분당)', '우리사주(분당)', '퇴직소득 과세', '퇴직소득 과세이연', '기타소득', '사업소득', '국내원천소득', '배당소득', '연말정산(분당)'],
                    '충무로': ['급여(충무로)', '연말정산(충무로)'],
                    '대덕': ['급여(대덕)', '연말정산(대덕)']
                }
                for workplace, items in ax_mappings.items():
                    self._aggregate_specific_items(workplace_summary[workplace], company_data, items)
        
        # 전체계 계산
        for workplace in workplace_summary:
            for data_type in ['인원', '소득세', '주민세']:
                total = 0
                for income_type in ['배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산']:
                    total += workplace_summary[workplace][income_type][data_type]
                workplace_summary[workplace]['전체계'][data_type] = total
        
        return workplace_summary

    def _aggregate_specific_items(self, workplace_data, company_data, item_names):
        """특정 항목들을 사업장 데이터에 집계"""
        for item_name in item_names:
            if item_name in company_data:
                item_data = company_data[item_name]
                # 소득 유형별로 분류해서 집계
                income_type = self._classify_income_type(item_name)
                workplace_data[income_type]['인원'] += item_data.get('인원', 0)
                workplace_data[income_type]['소득세'] += item_data.get('소득세', 0)
                workplace_data[income_type]['주민세'] += item_data.get('주민세', 0)

    def _classify_income_type(self, item_name):
        """항목명을 소득 유형으로 분류"""
        if '급여' in item_name or '일용' in item_name or '중도퇴사' in item_name or '우리사주' in item_name:
            return '근로소득'
        elif '퇴직' in item_name:
            return '퇴직소득'
        elif '배당' in item_name:
            return '배당소득'
        elif '사업' in item_name:
            return '사업소득'
        elif '기타' in item_name:
            return '기타소득'
        elif '법인원천' in item_name or '국내원천' in item_name:
            return '국내원천소득'
        elif '연말정산' in item_name:
            return '연말정산'
        else:
            return '기타소득'

    def _aggregate_company_to_workplace(self, workplace_data, company_data, company_type):
        """전체 회사 데이터를 사업장으로 집계"""
        for item_name, item_data in company_data.items():
            if '합계' not in item_name:
                income_type = self._classify_income_type(item_name)
                workplace_data[income_type]['인원'] += item_data.get('인원', 0)
                workplace_data[income_type]['소득세'] += item_data.get('소득세', 0)
                workplace_data[income_type]['주민세'] += item_data.get('주민세', 0)

    def create_excel_bytes(self, data_summary, company_type):
        """엑셀 파일을 바이트로 생성"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{COMPANY_CONFIGS[company_type]['name']} 보고서"

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            all_companies_data = {company_type: data_summary}
            self._create_main_report_sheet(ws, all_companies_data, thin_border)

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except Exception as e:
            logging.error(f"엑셀 생성 중 오류: {e}")
            return b''

    def create_combined_excel_bytes(self, all_companies_data):
        """통합 엑셀 파일을 바이트로 생성"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "1. 원천세 보고서"

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            self._create_main_report_sheet(ws, all_companies_data, thin_border)
            
            # 2번 시트: 신고양식 생성
            report_ws = wb.create_sheet(title="2. 신고양식(홈택스, 위택스)")
            
            # 홈택스용 데이터 계산
            companies_hometax_data = {}
            for company_type in [4, 1, 2, 3]:
                if company_type in all_companies_data:
                    companies_hometax_data[company_type] = self.extract_company_data_for_hometax(all_companies_data[company_type], company_type)

            # 총합계 계산
            total_hometax_data = self.calculate_total_hometax_data(companies_hometax_data)

            # 홈택스 신고 표 생성
            self._create_hometax_report(report_ws, total_hometax_data, companies_hometax_data, thin_border)

            # 통합 집계표 생성
            current_row = self._create_integrated_summary_table(report_ws, all_companies_data, thin_border, report_ws.max_row + 5)
            
            # 주민세 표 생성
            self._add_employee_resident_tax_table(report_ws, all_companies_data, thin_border)

            # 바이트로 변환
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except Exception as e:
            logging.error(f"통합 엑셀 생성 중 오류: {e}")
            return b''

    def _create_main_report_sheet(self, ws, all_companies_data, thin_border):
        """1번 시트 - 원천세 보고서 생성"""
        now = datetime.now()
        year = now.year
        month = now.month - 1
        
        if month == 0:
            month = 12
            year -= 1
        
        # 제목 추가
        title_text = f"{year}년 {month:02d}월 원천세 이행상황신고 보고서"
        title_cell = ws.cell(row=1, column=2, value=title_text)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
        
        headers = ['항목', '인원', '과세소득', '제출비과세', '총지급액', '소득세', '주민세']
        current_row = 3

        # 회사 순서: 4, 1, 2, 3
        for company_type in [4, 1, 2, 3]:
            if company_type not in all_companies_data:
                continue

            data_summary = all_companies_data[company_type]
            company_name = COMPANY_CONFIGS[company_type]['name']

            table_start_row = current_row

            # 회사 제목 추가
            title_cell = ws.cell(row=current_row, column=1, value=f"{company_name} 원천세 보고서")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            title_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            current_row += 1

            # 헤더 추가
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            current_row += 1

            # 데이터 순서 가져오기
            output_order = self._get_output_order_with_totals(company_type, data_summary)

            for item in output_order:
                if item in data_summary:
                    data = data_summary[item]
                    is_total = any(keyword in item for keyword in ['합계', '총합계', '최종'])

                    row_data = [
                        item,
                        format_number(data.get('인원', 0)),
                        format_number(data.get('과세소득', 0)) if item != '우리사주(분당)' else "",
                        format_number(data.get('제출비과세', 0)),
                        format_number(data.get('총지급액', 0)) if item != '우리사주(분당)' else "",
                        format_number(data.get('소득세', 0)),
                        format_number(data.get('주민세', 0))
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)

                        if is_total:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

                        if col > 1:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            if isinstance(value, str) and value.startswith('△'):
                                cell.font = Font(color="FF0000", bold=is_total)
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                    current_row += 1

            # 테두리 적용
            table_end_row = current_row - 1
            self._apply_table_borders(ws, table_start_row, table_end_row, 1, 7, thin_border)
            current_row += 2

        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15

    def _get_output_order_with_totals(self, company_type, data_summary):
        """회사별 출력 순서 반환"""
        orders = {
            1: [
                '급여(종로)', '사이닝보너스', '중도퇴사', '일용근로', '급여 총합계',
                '퇴직소득 과세이연', '퇴직소득 과세', '퇴직소득 총합계',
                '국내원천소득', '이자소득', '배당소득', '사업소득', '기타소득',
                '연말정산(합계)', '연말정산(분납액)', '연말정산(1월퇴사)', '연말정산(납부액)', '연말정산 전체 합계',
                '총합계'
            ],
            2: [
                '급여(INC 본점)', '급여(이천)', '일용근로소득', '급여 총합계',
                '중도퇴사연말정산(INC 본점)', '중도퇴사연말정산(이천)', '중도퇴사연말정산 총합계',
                '근로소득 총합계',
                '퇴직소득 과세', '퇴직소득 과세이연', '퇴직소득 총합계',
                '배당소득', '기타소득', '법인원천소득', '사업소득',
                '연말정산납부금액', '연말정산분납금액', '연말정산이천', '연말정산 총합계',
                '총합계'
            ],
            3: [
                '급여(MR 본점)', '급여(세종)', '급여(영주)', '급여(동탄)', '급여(상주)', '일용근로소득', '급여 총합계',
                '중도퇴사연말정산(MR 본점)', '중도퇴사연말정산(세종)', '중도퇴사연말정산(영주)', 
                '중도퇴사연말정산(동탄)', '중도퇴사연말정산(상주)', '중도퇴사연말정산 총합계',
                '근로소득 총합계',
                '퇴직소득 과세', '퇴직소득 과세이연(MR 본점)', '퇴직소득 과세이연(세종)', '퇴직소득 총합계',
                '법인원천소득', '배당소득', '이자소득', '기타소득', '사업소득',
                '연말정산(MR 본점)', '연말정산(세종)', '연말정산(영주)', '연말정산(동탄)', '연말정산(상주)', '연말정산 총합계',
                '총합계', '최종 총합계'
            ],
            4: [
                '급여(분당)', '급여(충무로)', '급여(대덕)', '일용직(3개월이상)', '일용근로소득', '급여 총합계',
                '중도퇴사연말정산(분당)', '중도퇴사연말정산 총합계',
                '우리사주(분당)', '우리사주 총합계',
                '근로소득 총합계',
                '퇴직소득 과세', '퇴직소득 과세이연', '퇴직소득 총합계',
                '국내원천소득', '배당소득', '기타소득', '사업소득',
                '연말정산(분당)', '연말정산(충무로)', '연말정산(대덕)', '연말정산 합계',
                '총합계', '최종 총합계'
            ]
        }

        order = orders.get(company_type, [])
        return [item for item in order if item in data_summary]

    def extract_company_data_for_hometax(self, company_data, company_type):
        """홈택스 신고용 회사별 데이터 추출"""
        result = {}
    
        def safe_get_data(key, default_keys=None):
            if default_keys is None:
                default_keys = ['인원', '인원수', '총지급액', '총지급액(과세소득+제출비과세)', '소득세']
            
            if key in company_data:
                data = company_data[key]
            else:
                data = {}
            
            safe_data = {
                '인원수': data.get('인원수', data.get('인원', 0)),
                '총지급액': data.get('총지급액', data.get('총지급액(과세소득+제출비과세)', 0)),
                '소득세': data.get('소득세', 0)
            }
            
            for k, v in safe_data.items():
                if v is None:
                    safe_data[k] = 0
            
            return safe_data
        
        # 애커튼(company_type=1) 특별 처리
        if company_type == 1:
            salary_jongro = safe_get_data('급여(종로)')
            signing_bonus = safe_get_data('사이닝보너스')
            
            result['간이세액'] = {
                '인원수': salary_jongro['인원수'] + signing_bonus['인원수'],
                '총지급액': salary_jongro['총지급액'] + signing_bonus['총지급액'],
                '소득세': salary_jongro['소득세'] + signing_bonus['소득세']
            }
            
            result['중도퇴사'] = safe_get_data('중도퇴사')
            result['일용근로소득'] = safe_get_data('일용근로')
            
        else:
            # 다른 회사들 기존 처리방식
            salary_total = safe_get_data('급여 총합계')
            daily_labor = safe_get_data('일용근로소득')
            
            result['간이세액'] = {
                '인원수': salary_total['인원수'] - daily_labor['인원수'],
                '총지급액': salary_total['총지급액'] - daily_labor['총지급액'],
                '소득세': salary_total['소득세'] - daily_labor['소득세']
            }
            
            result['중도퇴사'] = safe_get_data('중도퇴사연말정산 총합계')
            result['일용근로소득'] = daily_labor
            
        # 가감계1 (간이세액 + 중도퇴사 + 일용근로소득)
        result['가감계1'] = {
            '인원수': result['간이세액']['인원수'] + result['중도퇴사']['인원수'] + result['일용근로소득']['인원수'],
            '총지급액': result['간이세액']['총지급액'] + result['중도퇴사']['총지급액'] + result['일용근로소득']['총지급액'],
            '소득세': result['간이세액']['소득세'] + result['중도퇴사']['소득세'] + result['일용근로소득']['소득세']
        }
        
        # 퇴직소득
        result['퇴직소득과세이연'] = safe_get_data('퇴직소득 과세이연')
        result['퇴직소득과세'] = safe_get_data('퇴직소득 과세')
        
        # 가감계2 (퇴직소득 합계)
        result['가감계2'] = {
            '인원수': result['퇴직소득과세이연']['인원수'] + result['퇴직소득과세']['인원수'],
            '총지급액': result['퇴직소득과세이연']['총지급액'] + result['퇴직소득과세']['총지급액'],
            '소득세': result['퇴직소득과세이연']['소득세'] + result['퇴직소득과세']['소득세']
        }
    
        # 개별 소득 항목들
        result['사업소득'] = safe_get_data('사업소득')
        result['기타소득'] = safe_get_data('기타소득')
        
        # 법인원천 (법인원천소득 또는 국내원천소득)
        legal_source = safe_get_data('법인원천소득') if '법인원천소득' in company_data else safe_get_data('국내원천소득')
        result['법인원천'] = legal_source
        
        result['배당소득'] = safe_get_data('배당소득')
        
        # 연말정산
        yearend_data = {'인원수': 0, '총지급액': 0, '소득세': 0}
        for key in company_data.keys():
            if ('연말정산' in key and '합계' in key and '중도퇴사' not in key) or key == '연말정산 전체 합계':
                yearend_data = safe_get_data(key)
                break
        result['연말정산'] = yearend_data
        
        # 이자소득 (0 고정)
        result['이자소득'] = {'인원수': 0, '총지급액': 0, '소득세': 0}
        
        # 총합계
        total_personnel = (result['가감계1']['인원수'] + result['가감계2']['인원수'] + 
                          result['사업소득']['인원수'] + result['기타소득']['인원수'] + 
                          result['법인원천']['인원수'] + result['배당소득']['인원수'] + 
                          result['연말정산']['인원수'] + result['이자소득']['인원수'])
        
        total_amount = (result['가감계1']['총지급액'] + result['가감계2']['총지급액'] + 
                       result['사업소득']['총지급액'] + result['기타소득']['총지급액'] + 
                       result['법인원천']['총지급액'] + result['배당소득']['총지급액'] + 
                       result['연말정산']['총지급액'] + result['이자소득']['총지급액'])
        
        total_tax = (result['가감계1']['소득세'] + result['가감계2']['소득세'] + 
                    result['사업소득']['소득세'] + result['기타소득']['소득세'] + 
                    result['법인원천']['소득세'] + result['배당소득']['소득세'] + 
                    result['연말정산']['소득세'] + result['이자소득']['소득세'])
        
        result['총합계'] = {'인원수': total_personnel, '총지급액': total_amount, '소득세': total_tax}
    
        # 고정값들
        result['농특세'] = {'인원수': 0, '총지급액': 0, '소득세': 0}
        result['총합계2'] = result['총합계'].copy()
        result['전월미환급세액'] = {'인원수': 0, '총지급액': 0, '소득세': 0}
        result['최종납부할세액'] = result['총합계'].copy()
    
        return result

    def calculate_total_hometax_data(self, companies_hometax_data):
        """4개 회사 홈택스 데이터 총합계 계산"""
        total_data = {}
        
        for item in HOMETAX_ITEMS:
            total_personnel = sum(companies_hometax_data[comp][item]['인원수'] for comp in [4, 1, 2, 3] if comp in companies_hometax_data)
            total_amount = sum(companies_hometax_data[comp][item].get('총지급액', 0) for comp in [4, 1, 2, 3] if comp in companies_hometax_data)
            total_tax = sum(companies_hometax_data[comp][item]['소득세'] for comp in [4, 1, 2, 3] if comp in companies_hometax_data)
            
            total_data[item] = {
                '인원수': total_personnel,
                '총지급액': total_amount,
                '소득세': total_tax
            }
        
        return total_data

    def _create_hometax_report(self, report_ws, total_hometax_data, companies_hometax_data, thin_border):
        """홈택스 신고 표 생성"""
        # 전체 시트 제목
        title_cell = report_ws.cell(row=1, column=1, value="홈택스/위택스 각 세목별 신고 납부액 계산")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        report_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)

        # 홈택스 신고 표 제목
        hometax_title = report_ws.cell(row=3, column=1, value="홈택스 신고")
        hometax_title.font = Font(bold=True, size=14)
        hometax_title.alignment = Alignment(horizontal='center', vertical='center')
        hometax_title.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        report_ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

        # 홈택스 신고 표 헤더
        headers = ['구분', '인원', '총지급액', '소득세']
        for col, header in enumerate(headers, 1):
            cell = report_ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # 홈택스 신고 표 데이터 행들
        current_row = 5
        for item in HOMETAX_ITEMS:
            cell = report_ws.cell(row=current_row, column=1, value=item)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if item in ['가감계1', '가감계2', '총합계', '총합계2', '최종납부할세액']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

            data = total_hometax_data.get(item, {'인원수': 0, '총지급액': 0, '소득세': 0})
            
            values = [
                format_number(data.get('인원수', 0)),
                format_number(data.get('총지급액', 0)),
                format_number(data.get('소득세', 0))
            ]
            
            for col, value in enumerate(values, 2):
                cell = report_ws.cell(row=current_row, column=col, value=value)
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
                if isinstance(value, str) and value.startswith('△'):
                    if item in ['가감계1', '가감계2', '총합계', '총합계2', '최종납부할세액']:
                        cell.font = Font(color="FF0000", bold=True)
                    else:
                        cell.font = Font(color="FF0000")
                else:
                    if item in ['가감계1', '가감계2', '총합계', '총합계2', '최종납부할세액']:
                        cell.font = Font(bold=True)
                
                if item in ['가감계1', '가감계2', '총합계', '총합계2', '최종납부할세액']:
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            current_row += 1

        self._apply_hometax_table_borders(report_ws, 3, current_row-1, 1, 4)

        report_ws.column_dimensions['A'].width = 20
        report_ws.column_dimensions['B'].width = 12
        report_ws.column_dimensions['C'].width = 15
        report_ws.column_dimensions['D'].width = 15

    def _create_integrated_summary_table(self, report_ws, all_companies_data, thin_border, start_row):
        """통합 집계표 생성"""
        title_cell = report_ws.cell(row=start_row, column=1, value="위택스 신고 내역")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        report_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=25)
        
        start_row += 1
        
        workplace_summary = self._aggregate_workplace_data(all_companies_data)
        
        headers = ['사업장', '배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산', '전체계']
        
        col_idx = 1
        for header in headers:
            if header == '사업장':
                cell = report_ws.cell(row=start_row, column=col_idx, value=header)
                report_ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=start_row+1, end_column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                col_idx += 1
            else:
                cell = report_ws.cell(row=start_row, column=col_idx, value=header)
                report_ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=start_row, end_column=col_idx+2)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                col_idx += 3
        
        col_idx = 2
        for income_type in ['배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산', '전체계']:
            for sub_header in ['인원', '소득세', '주민세']:
                cell = report_ws.cell(row=start_row+1, column=col_idx, value=sub_header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                col_idx += 1
        
        data_start_row = start_row + 2
        total_data = {income_type: {'인원': 0, '소득세': 0, '주민세': 0} for income_type in ['배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산', '전체계']}
        
        for row_idx, workplace in enumerate(WORKPLACE_ORDER):
            if workplace not in workplace_summary:
                continue
                
            current_row = data_start_row + row_idx
            
            cell = report_ws.cell(row=current_row, column=1, value=workplace)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            col_idx = 2
            for income_type in ['배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산', '전체계']:
                data = workplace_summary[workplace].get(income_type, {'인원': 0, '소득세': 0, '주민세': 0})
                
                for data_type in ['인원', '소득세', '주민세']:
                    value = data.get(data_type, 0)
                    cell = report_ws.cell(row=current_row, column=col_idx, value=format_number(value))
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    if data_type not in total_data[income_type]:
                        total_data[income_type][data_type] = 0
                    total_data[income_type][data_type] += value
                    col_idx += 1
        
        total_row = data_start_row + len([wp for wp in WORKPLACE_ORDER if wp in workplace_summary])
        
        cell = report_ws.cell(row=total_row, column=1, value="합계")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        col_idx = 2
        for income_type in ['배당소득', '사업소득', '근로소득', '기타소득', '퇴직소득', '국내원천소득', '연말정산', '전체계']:
            for data_type in ['인원', '소득세', '주민세']:
                value = total_data[income_type].get(data_type, 0)
                cell = report_ws.cell(row=total_row, column=col_idx, value=format_number(value))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                col_idx += 1
        
        self._apply_table_borders(report_ws, start_row - 1, total_row, 1, col_idx-1, thin_border)

        report_ws.column_dimensions['A'].width = 15
        for col in range(2, 25):
            col_letter = get_column_letter(col)
            report_ws.column_dimensions[col_letter].width = 12

        return total_row + 3

    def _add_employee_resident_tax_table(self, report_ws, all_companies_data, thin_border):
        """주민세 종업원분 내역 표를 추가하는 함수"""
        start_row = report_ws.max_row + 3
        
        companies = {
            2: all_companies_data.get(2, {}),
            4: all_companies_data.get(4, {}),  
            3: all_companies_data.get(3, {}),
            1: all_companies_data.get(1, {})
        }
        
        workplace_results = {}
        
        calculations = {
            2: [
                {'workplace': 'INC본점', 'items': ['급여(INC 본점)', '일용근로소득']},
                {'workplace': '이천', 'items': ['급여(이천)']}
            ],
            4: [
                {'workplace': '분당', 'items': ['급여(분당)', '일용직(3개월이상)', '일용근로소득']},
                {'workplace': '충무로', 'items': ['급여(충무로)']},
                {'workplace': '대덕', 'items': ['급여(대덕)']}
            ],
            3: [
                {'workplace': 'MR본점', 'items': ['급여(MR 본점)', '일용근로소득']},
                {'workplace': '세종', 'items': ['급여(세종)']},
                {'workplace': '동탄', 'items': ['급여(동탄)']},
                {'workplace': '상주', 'items': ['급여(상주)']}
            ],
            1: [
                {'workplace': '애커튼', 'items': ['급여(종로)', '일용근로']}
            ]
        }
        
        for company_type, calc_info in calculations.items():
            if company_type not in companies or not companies[company_type]:
                continue
                
            company_data = companies[company_type]
            
            for calc in calc_info:
                total_people = sum(company_data.get(item, {}).get('인원', 0) for item in calc['items'])
                total_income = sum(company_data.get(item, {}).get('과세소득', 0) for item in calc['items'])
                resident_tax = math.floor(total_income * 0.005 / 10) * 10
                
                workplace_results[calc['workplace']] = {
                    '종업원수': total_people,
                    '계산근거': total_income,
                    '주민세종업원분': resident_tax
                }
        
        total_people = sum(data['종업원수'] for data in workplace_results.values())
        total_income = sum(data['계산근거'] for data in workplace_results.values())
        total_resident_tax = sum(data['주민세종업원분'] for data in workplace_results.values())
        
        title_cell = report_ws.cell(row=start_row, column=1, value="주민세 종업원분 내역")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        report_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        
        header_row = start_row + 1
        headers = ['사업장', '종업원수', '계산근거(과세소득×0.5%)', '주민세종업원분']
        
        for col, header in enumerate(headers, 1):
            cell = report_ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        data_start_row = start_row + 2
        workplaces = ['INC본점', '분당', '충무로', 'MR본점', '애커튼']
        
        for row_idx, workplace in enumerate(workplaces):
            current_row = data_start_row + row_idx
            
            cell = report_ws.cell(row=current_row, column=1, value=workplace)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            if workplace in workplace_results:
                data = workplace_results[workplace]
                values = [data['종업원수'], data['계산근거'], data['주민세종업원분']]
                for col, value in enumerate(values, 2):
                    cell = report_ws.cell(row=current_row, column=col, value=format_number(value))
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                for col in range(2, 5):
                    cell = report_ws.cell(row=current_row, column=col, value="")
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        total_row = data_start_row + len(workplaces)
        
        cell = report_ws.cell(row=total_row, column=1, value="합계")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        totals = [total_people, total_income, total_resident_tax]
        for col, total_value in enumerate(totals, 2):
            cell = report_ws.cell(row=total_row, column=col, value=format_number(total_value))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        end_row = total_row
        self._apply_table_borders(report_ws, start_row, end_row, 1, 4, thin_border)
        
        column_widths = [15, 12, 20, 15]
        for col, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col)
            report_ws.column_dimensions[col_letter].width = width

    def _apply_table_borders(self, ws, start_row, end_row, start_col, end_col, border_style):
        """표에 테두리 적용"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border_style
                
                border_styles = {}
                if row == start_row:
                    border_styles['top'] = Side(style='thick')
                if row == end_row:
                    border_styles['bottom'] = Side(style='thick')
                if col == start_col:
                    border_styles['left'] = Side(style='thick')
                if col == end_col:
                    border_styles['right'] = Side(style='thick')
                
                if border_styles:
                    new_border = Border(
                        left=border_styles.get('left', cell.border.left),
                        right=border_styles.get('right', cell.border.right),
                        top=border_styles.get('top', cell.border.top),
                        bottom=border_styles.get('bottom', cell.border.bottom)
                    )
                    cell.border = new_border

    def _apply_hometax_table_borders(self, ws, start_row, end_row, start_col, end_col):
        """홈택스 표에 테두리 적용"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
            
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                    
                border_styles = {}
                if row == start_row:
                    border_styles['top'] = Side(style='thick')
                if row == end_row:
                    border_styles['bottom'] = Side(style='thick')
                if col == start_col:
                    border_styles['left'] = Side(style='thick')
                if col == end_col:
                    border_styles['right'] = Side(style='thick')
                    
                if border_styles:
                    new_border = Border(
                        left=border_styles.get('left', cell.border.left),
                        right=border_styles.get('right', cell.border.right),
                        top=border_styles.get('top', cell.border.top),
                        bottom=border_styles.get('bottom', cell.border.bottom)
                    )
                    cell.border = new_border